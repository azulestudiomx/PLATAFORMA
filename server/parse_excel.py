import json
import os
import openpyxl
from datetime import datetime

# Input and output paths
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'ESTRUCTURA  ESTATAL .xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'padron_campeche.json')

# Standardized sheet name mappings (or fallback to original names)
SHEETS = ['CAMP', 'CAND', 'HECE', 'CARM', 'CALA', 'CHAM', 'PALI', 'TENA', 'CALK', 'DZIT', 'SEYB', 'ESC']

# Normalized mapping from header strings to Person fields
HEADER_MAP = {
    'nombre': 'name',
    'nombre completo': 'name',
    'dirección': 'address',
    'direccion': 'address',
    'municipio': 'municipio',
    'distrito': 'distrito',
    'distrito local': 'distrito',
    'distrito local ': 'distrito',
    'sección': 'seccion',
    'seccion': 'seccion',
    'celular': 'phone',
    'telefono': 'phone',
    'teléfono': 'phone',
    'correo': 'email',
    'correo ': 'email',
    'correo electrónico': 'email',
    'email': 'email',
    'cumpleaños': 'birthday',
    'cumpleaños ': 'birthday',
    'cumpleanos': 'birthday',
    'fecha de nacimiento': 'birthday'
}

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    
    # Remove floating point artifacts e.g. "9811685660.0" or "123.0"
    if s.endswith('.0') and s[:-2].replace('-', '').isdigit():
        s = s[:-2]
    return s

def format_birthday(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # If it is a datetime represented as string with hours like "1990-03-15 00:00:00"
    if ' ' in s and s.split()[0].replace('-', '').isdigit() and len(s.split()[0]) == 10:
        return s.split()[0]
    return s

def title_case_name(name):
    """Capitalizes names properly in Spanish (e.g. Juan De La Cruz)."""
    if not name:
        return name
    exceptions = {'de', 'del', 'la', 'las', 'los', 'y', 'e', 'a'}
    words = name.lower().split()
    result = []
    for i, word in enumerate(words):
        if i == 0 or word not in exceptions:
            result.append(word.capitalize())
        else:
            result.append(word)
    return ' '.join(result)

def main():
    print(f"📖 Iniciando procesamiento de Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: El archivo no existe en {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    records = []
    total_raw_rows = 0

    # Process sheet by sheet
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEETS:
            print(f"⚠️  Saltando hoja no configurada: {sheet_name}")
            continue

        sheet = wb[sheet_name]
        print(f"   📂 Procesando hoja: {sheet_name}")
        
        # 1. Locate headers row
        header_row_index = None
        col_mapping = {} # maps col_idx -> person_field
        
        # We search first 10 rows to detect the header
        rows_iterator = sheet.iter_rows(values_only=True)
        rows_list = []
        for r_idx in range(1, 12):
            try:
                row = next(rows_iterator)
                rows_list.append((r_idx, row))
            except StopIteration:
                break
        
        for idx, row in rows_list:
            if not row:
                continue
            # Check if this row contains 'nombre' or 'nombre completo' or similar header keywords
            row_str_vals = [str(val).strip().lower() for val in row if val is not None]
            if any(k in row_str_vals for k in ['nombre', 'nombre completo', 'no.', 'no']):
                header_row_index = idx
                # Map columns dynamically
                for col_idx, cell_val in enumerate(row):
                    if cell_val is not None:
                        clean_header = str(cell_val).strip().lower()
                        if clean_header in HEADER_MAP:
                            col_mapping[col_idx] = HEADER_MAP[clean_header]
                break

        if header_row_index is None:
            print(f"   ⚠️  No se encontró fila de encabezado en la hoja {sheet_name}. Saltando.")
            continue

        print(f"      Headers encontrados en fila {header_row_index}. Columnas mapeadas: {col_mapping}")
        
        # Re-iterate or continue iteration from the row after headers
        # Since iter_rows is read-only, we can just instantiate a fresh iterator and skip first header_row_index rows
        sheet_rows = sheet.iter_rows(values_only=True)
        for _ in range(header_row_index):
            next(sheet_rows)
            
        # Parse data rows
        sheet_records_count = 0
        for row_idx, row in enumerate(sheet_rows, start=header_row_index + 1):
            total_raw_rows += 1
            
            # Map row columns to fields
            data = {}
            for col_idx, field in col_mapping.items():
                if col_idx < len(row):
                    data[field] = row[col_idx]
            
            # Extract and validate name
            raw_name = clean_str(data.get('name'))
            if not raw_name or raw_name.lower() in ('nombre', 'no.', 'no', 'estructura', ''):
                continue # Skip residual headers or blank rows

            # Clean and normalize fields
            nombre_limpio = title_case_name(raw_name)
            phone = clean_str(data.get('phone'))
            address = clean_str(data.get('address'))
            
            # Default or fallback municipality based on sheet name if not in cell
            municipio = clean_str(data.get('municipio'))
            if not municipio:
                # Map sheet codes to full names
                municipio_fallbacks = {
                    'CAMP': 'Campeche', 'CAND': 'Candelaria', 'HECE': 'Hecelchakán',
                    'CARM': 'Carmen', 'CALA': 'Calakmul', 'CHAM': 'Champotón',
                    'PALI': 'Palizada', 'TENA': 'Tenabo', 'CALK': 'Calkiní',
                    'DZIT': 'Dzitbalché', 'SEYB': 'Seybaplaya', 'ESC': 'Escárcega'
                }
                municipio = municipio_fallbacks.get(sheet_name, sheet_name)
            
            distrito = clean_str(data.get('distrito'))
            seccion = clean_str(data.get('seccion'))
            email = clean_str(data.get('email'))
            birthday = format_birthday(data.get('birthday'))
            
            # Create standard record
            record = {
                "name": nombre_limpio,
                "phone": phone,
                "address": address,
                "municipio": municipio,
                "localidad": "",
                "distrito": distrito,
                "zona": "",
                "seccion": seccion,
                "email": email,
                "birthday": birthday,
                "synced": 1
            }
            records.append(record)
            sheet_records_count += 1
            
        print(f"      ✅ Hoja {sheet_name} procesada. Registros válidos: {sheet_records_count}")

    # Generate sequential unique INEs
    for idx, r in enumerate(records, start=1):
        r["ine"] = f"CAMP-{str(idx).zfill(4)}"

    print(f"\n📊 RESUMEN DE PROCESAMIENTO:")
    print(f"   Total de registros válidos: {len(records)}")
    
    # Save JSON output
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
        
    print(f"✅ JSON exportado correctamente a: {OUTPUT_PATH}")
    print("\nPrimeros 5 registros de muestra:")
    for r in records[:5]:
        print(f"  {r['ine']}: {r['name']} | Cel: {r['phone']} | Correo: {r['email']} | Cumple: {r['birthday']} | Secc: {r['seccion']}")

if __name__ == '__main__':
    main()
